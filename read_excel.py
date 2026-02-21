# -*- coding: utf-8 -*-
import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'D:\outbound-management 250728 - 백업(여기로돌아오기) - 복사본\d-care-console\docs\OpenAPI 통화내역_20260105.xlsx')
ws = wb.active

print(f'총 {ws.max_row}행, {ws.max_column}열')
print('---')

for row in ws.iter_rows(max_row=60):
    vals = [str(cell.value or '') for cell in row]
    if any(vals):
        print(' | '.join(vals))
